#!/usr/bin/env python3
"""
render_excel.py -- render plan_model.json into an interactive Excel workbook.

Usage:
    python render_excel.py plan_model.json --out financial_plan.xlsx

Reads the plan data model (see references/plan-model.md for the schema) and
produces a multi-sheet workbook: Dashboard, Assumptions, Net Worth, Cash Flow,
Debt, Retirement, Transactions, Source Manifest.

This script does NOT recompute analysis -- it only renders analysis.* fields
that scripts/projections.py (or the calling session) already filled in. If a
number looks wrong, fix plan_model.json and re-render.
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl  # noqa: F401
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    print(
        "openpyxl is required but not installed for this Python interpreter.\n"
        "Install it with:\n"
        r"    C:\AI-Shared\python.exe -m pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)


# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

TITLE_FONT = Font(size=16, bold=True, color="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
SUBHEAD_FONT = Font(size=12, bold=True, color="1F2937")
BIG_LABEL_FONT = Font(size=11, bold=True, color="374151")
BIG_VALUE_FONT = Font(size=20, bold=True, color="0F766E")
NEGATIVE_VALUE_FONT = Font(size=20, bold=True, color="B91C1C")
INPUT_FILL = PatternFill("solid", fgColor="FEF9C3")  # editable cells
INPUT_FONT = Font(color="92400E", bold=False)
DISCLAIMER_FONT = Font(size=9, italic=True, color="6B7280")
DISCLAIMER_FILL = PatternFill("solid", fgColor="F3F4F6")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
CARD_FILL = PatternFill("solid", fgColor="F0FDFA")
CURRENCY_FMT = '#,##0.00_);[RED](#,##0.00)'
PCT_FMT = '0.0"%"'


def money(v):
    return v if isinstance(v, (int, float)) else 0.0


def safe_get(d, *path, default=None):
    cur = d
    for p in path:
        if cur is None:
            return default
        if isinstance(cur, dict):
            cur = cur.get(p)
        elif isinstance(cur, list) and isinstance(p, int):
            cur = cur[p] if -len(cur) <= p < len(cur) else None
        else:
            return default
    return default if cur is None else cur


def autosize(ws: Worksheet, min_width=10, max_width=48):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column_letter
            length = len(str(cell.value))
            widths[col] = max(widths.get(col, 0), length)
    for col, length in widths.items():
        ws.column_dimensions[col].width = max(min_width, min(max_width, length + 2))


def write_title(ws, text, row=1, col=1, span=6):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = TITLE_FONT
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    return row + 2


def write_disclaimer(ws, disclaimer, row, col=1, span=8):
    cell = ws.cell(row=row, column=col, value=f"Disclaimer: {disclaimer}")
    cell.font = DISCLAIMER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col + span - 1)
    for r in range(row, row + 2):
        for c in range(col, col + span):
            ws.cell(row=r, column=c).fill = DISCLAIMER_FILL
    ws.row_dimensions[row].height = 30
    return row + 3


def write_table(ws, start_row, start_col, headers, rows, header_fill=True):
    """Write a simple header + rows table; returns the row after the table."""
    r = start_row
    for j, h in enumerate(headers):
        c = ws.cell(row=r, column=start_col + j, value=h)
        c.font = HEADER_FONT
        if header_fill:
            c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")
    r += 1
    first_data_row = r
    for row_vals in rows:
        for j, v in enumerate(row_vals):
            c = ws.cell(row=r, column=start_col + j, value=v)
            c.border = THIN_BORDER
            if isinstance(v, (int, float)):
                c.number_format = CURRENCY_FMT
        r += 1
    return first_data_row, r  # (first data row, next free row)


def input_cell(ws, row, col, value, number_format=None, label=None):
    if label is not None:
        lc = ws.cell(row=row, column=col, value=label)
        lc.font = BIG_LABEL_FONT
        col += 1
    c = ws.cell(row=row, column=col, value=value)
    c.fill = INPUT_FILL
    c.font = INPUT_FONT
    c.border = THIN_BORDER
    if number_format:
        c.number_format = number_format
    return c


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_dashboard(wb, model):
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    row = write_title(ws, "Financial Plan Dashboard", row=1, span=8)
    gen = model.get("generated_at", "unknown")
    ws.cell(row=row, column=1, value=f"Generated: {gen}    Currency: {model.get('currency', 'USD')}")
    ws.cell(row=row, column=1).font = Font(italic=True, color="6B7280")
    row += 2

    nw = safe_get(model, "analysis", "net_worth", default={}) or {}
    cf = safe_get(model, "analysis", "cash_flow", default={}) or {}
    ef = safe_get(model, "analysis", "emergency_fund", default={}) or {}
    debt = safe_get(model, "analysis", "debt", default={}) or {}

    # --- Big labeled KPI cells (a "card" row) ---
    cards = [
        ("Net Worth", money(nw.get("net")), True),
        ("Savings Rate", cf.get("savings_rate_pct"), False, "pct"),
        ("Emergency Runway (months)", ef.get("runway_months"), False),
        ("Total Debt", money(debt.get("total")), True),
    ]
    card_col = 1
    card_row_label = row
    card_row_value = row + 1
    for card in cards:
        name = card[0]
        value = card[1]
        is_money = card[2]
        kind = card[3] if len(card) > 3 else ("money" if is_money else "num")

        lbl = ws.cell(row=card_row_label, column=card_col, value=name)
        lbl.font = BIG_LABEL_FONT
        lbl.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=card_row_label, start_column=card_col,
                        end_row=card_row_label, end_column=card_col + 1)

        vcell = ws.cell(row=card_row_value, column=card_col, value=value if value is not None else "n/a")
        vcell.alignment = Alignment(horizontal="center")
        if isinstance(value, (int, float)):
            if kind == "money":
                vcell.number_format = '"$"#,##0'
            elif kind == "pct":
                vcell.number_format = PCT_FMT
            vcell.font = BIG_VALUE_FONT if (value is None or value >= 0) else NEGATIVE_VALUE_FONT
        else:
            vcell.font = BIG_VALUE_FONT
        ws.merge_cells(start_row=card_row_value, start_column=card_col,
                        end_row=card_row_value, end_column=card_col + 1)
        for r in (card_row_label, card_row_value):
            for c in range(card_col, card_col + 2):
                ws.cell(row=r, column=c).fill = CARD_FILL
                ws.cell(row=r, column=c).border = THIN_BORDER
        card_col += 2

    row = card_row_value + 2

    # --- Debt-free date / goal progress ---
    avalanche = debt.get("avalanche", {}) or {}
    months_to_free = avalanche.get("months_to_free")
    debt_free_note = "n/a — no debt data" if not months_to_free else f"~{months_to_free} months (avalanche order)"
    ws.cell(row=row, column=1, value="Debt-free date (est.)").font = BIG_LABEL_FONT
    ws.cell(row=row, column=2, value=debt_free_note).font = Font(bold=True, color="0F766E")
    row += 2

    goals = safe_get(model, "profile", "goals", default=[]) or []
    if goals:
        ws.cell(row=row, column=1, value="Goal Progress").font = SUBHEAD_FONT
        row += 1
        headers = ["Goal", "Current", "Target", "% Complete", "Target Date"]
        rows = []
        for g in goals:
            cur = money(g.get("current_amount"))
            tgt = money(g.get("target_amount")) or 0
            pct = round((cur / tgt) * 100, 1) if tgt else None
            rows.append([g.get("name"), cur, tgt, pct, g.get("target_date")])
        first, row = write_table(ws, row, 1, headers, rows)
        # percent format + simple text progress bar via conditional-free approach
        for i in range(first, row):
            c = ws.cell(row=i, column=4)
            if isinstance(c.value, (int, float)):
                c.number_format = PCT_FMT
        row += 1
    else:
        ws.cell(row=row, column=1, value="Goal Progress: no goals recorded").font = SUBHEAD_FONT
        row += 2

    # --- Charts ---
    chart_anchor_row = row
    by_cat = cf.get("by_category") or {}
    if by_cat:
        cat_start = row
        ws.cell(row=row, column=1, value="Spending by Category").font = SUBHEAD_FONT
        row += 1
        cat_header_row = row
        headers = ["Category", "Monthly Avg"]
        rows = [[k, money(v)] for k, v in by_cat.items()]
        first, row = write_table(ws, row, 1, headers, rows)

        bar = BarChart()
        bar.title = "Spending by Category"
        bar.style = 10
        bar.y_axis.title = "Monthly $"
        data = Reference(ws, min_col=2, min_row=cat_header_row, max_row=row - 1)
        cats = Reference(ws, min_col=1, min_row=first, max_row=row - 1)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.width = 16
        bar.height = 8
        ws.add_chart(bar, f"F{cat_start}")
        row += 1

    proj = safe_get(model, "analysis", "retirement", "projection_series", default=[]) or []
    if proj:
        ret_start = row
        ws.cell(row=row, column=1, value="Retirement Projection (Best/Base/Worst)").font = SUBHEAD_FONT
        row += 1
        ret_header_row = row
        headers = ["Age", "Best", "Base", "Worst"]
        rows = [[p.get("age"), money(p.get("best")), money(p.get("base")), money(p.get("worst"))] for p in proj]
        first, row = write_table(ws, row, 1, headers, rows)

        line = LineChart()
        line.title = "Retirement Projection"
        line.style = 12
        line.y_axis.title = "Projected value ($)"
        line.x_axis.title = "Age"
        data = Reference(ws, min_col=2, max_col=4, min_row=ret_header_row, max_row=row - 1)
        cats = Reference(ws, min_col=1, min_row=first, max_row=row - 1)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.width = 16
        line.height = 8
        ws.add_chart(line, f"F{ret_start}")
        row += 1

    row += 2
    row = write_disclaimer(ws, model.get("disclaimer", ""), row)
    autosize(ws)
    return ws


def build_assumptions(wb, model):
    ws = wb.create_sheet("Assumptions")
    row = write_title(ws, "Assumptions (editable)", span=4)
    ws.cell(row=row, column=1,
            value="Yellow cells are inputs. Change them and re-run the refresh path to "
                  "recompute the plan; some cells on this sheet recompute live via formula.")
    ws.cell(row=row, column=1).font = Font(italic=True, color="6B7280")
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    a = model.get("assumptions", {}) or {}
    ef = safe_get(model, "analysis", "emergency_fund", default={}) or {}
    cf = safe_get(model, "analysis", "cash_flow", default={}) or {}

    fields = [
        ("expected_return_pct", "Expected return (%)", PCT_FMT),
        ("inflation_pct", "Inflation (%)", PCT_FMT),
        ("retirement_age", "Retirement age", "0"),
        ("current_age", "Current age", "0"),
        ("monthly_contribution", "Monthly contribution ($)", CURRENCY_FMT),
        ("emergency_fund_months", "Emergency fund target (months)", "0"),
    ]

    label_col, value_col = 1, 2
    cell_refs = {}
    for key, label, fmt in fields:
        ws.cell(row=row, column=label_col, value=label).font = BIG_LABEL_FONT
        c = input_cell(ws, row, value_col, a.get(key))
        c.number_format = fmt
        cell_refs[key] = f"{c.coordinate}"
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Notes").font = BIG_LABEL_FONT
    note_cell = input_cell(ws, row, 2, a.get("notes", ""))
    note_cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
    row += 3

    # --- Live formula: emergency fund target = months x average monthly expenses ---
    ws.cell(row=row, column=1, value="Live formulas (recompute automatically)").font = SUBHEAD_FONT
    row += 1
    ws.cell(row=row, column=1, value="Avg monthly expenses (from Cash Flow sheet)").font = BIG_LABEL_FONT
    exp_cell = ws.cell(row=row, column=2, value=money(cf.get("monthly_expenses_avg")))
    exp_cell.number_format = CURRENCY_FMT
    exp_cell.border = THIN_BORDER
    row += 1

    ws.cell(row=row, column=1, value="Emergency fund target ($) = months x avg expenses").font = BIG_LABEL_FONT
    months_ref = cell_refs.get("emergency_fund_months")
    formula = f"={months_ref}*{exp_cell.coordinate}" if months_ref else None
    target_cell = ws.cell(row=row, column=2, value=formula if formula else money(ef.get("target_amount")))
    target_cell.number_format = CURRENCY_FMT
    target_cell.font = Font(bold=True, color="0F766E")
    target_cell.border = THIN_BORDER
    row += 2

    ws.cell(row=row, column=1,
            value="Deeper scenario math (debt payoff order, retirement scenarios) is "
                  "computed by scripts/projections.py and shown as static values on their "
                  "sheets — recompute via refresh, not by editing formulas here.")
    ws.cell(row=row, column=1).font = Font(italic=True, color="B45309")
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    autosize(ws)
    return ws, cell_refs


def build_net_worth(wb, model):
    ws = wb.create_sheet("Net Worth")
    row = write_title(ws, "Net Worth", span=5)

    nw = safe_get(model, "analysis", "net_worth", default={}) or {}
    ws.cell(row=row, column=1, value="Assets").font = BIG_LABEL_FONT
    ws.cell(row=row, column=2, value=money(nw.get("assets"))).number_format = CURRENCY_FMT
    row += 1
    ws.cell(row=row, column=1, value="Liabilities").font = BIG_LABEL_FONT
    ws.cell(row=row, column=2, value=money(nw.get("liabilities"))).number_format = CURRENCY_FMT
    row += 1
    ws.cell(row=row, column=1, value="Net Worth").font = BIG_LABEL_FONT
    net_cell = ws.cell(row=row, column=2, value=money(nw.get("net")))
    net_cell.number_format = CURRENCY_FMT
    net_cell.font = Font(bold=True, color="0F766E")
    row += 3

    accounts = model.get("accounts", []) or []
    ws.cell(row=row, column=1, value="Accounts").font = SUBHEAD_FONT
    row += 1
    headers = ["Name", "Institution", "Mask", "Type", "Balance", "As Of", "Liability?"]
    rows = [
        [acc.get("name"), acc.get("institution"), acc.get("mask"), acc.get("type"),
         money(acc.get("balance")), acc.get("as_of"), "Yes" if acc.get("is_liability") else "No"]
        for acc in accounts
    ]
    _, row = write_table(ws, row, 1, headers, rows)
    row += 2

    trend = nw.get("trend") or []
    if trend:
        ws.cell(row=row, column=1, value="Net Worth Trend").font = SUBHEAD_FONT
        row += 1
        t_header_row = row
        headers = ["Date", "Net"]
        rows = [[t.get("date"), money(t.get("net"))] for t in trend]
        first, row = write_table(ws, row, 1, headers, rows)
        if len(rows) > 1:
            line = LineChart()
            line.title = "Net Worth Trend"
            line.y_axis.title = "Net ($)"
            data = Reference(ws, min_col=2, min_row=t_header_row, max_row=row - 1)
            cats = Reference(ws, min_col=1, min_row=first, max_row=row - 1)
            line.add_data(data, titles_from_data=True)
            line.set_categories(cats)
            line.width = 14
            line.height = 7
            ws.add_chart(line, f"D{t_header_row}")

    autosize(ws)
    return ws


def build_cash_flow(wb, model):
    ws = wb.create_sheet("Cash Flow")
    row = write_title(ws, "Cash Flow / Budget", span=4)

    cf = safe_get(model, "analysis", "cash_flow", default={}) or {}
    kv = [
        ("Monthly income (avg)", money(cf.get("monthly_income_avg"))),
        ("Monthly expenses (avg)", money(cf.get("monthly_expenses_avg"))),
        ("Surplus / (deficit)", money(cf.get("surplus"))),
        ("Savings rate (%)", cf.get("savings_rate_pct")),
    ]
    for label, val in kv:
        ws.cell(row=row, column=1, value=label).font = BIG_LABEL_FONT
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = PCT_FMT if "%" in label else CURRENCY_FMT
        row += 1
    row += 1

    by_cat = cf.get("by_category") or {}
    ws.cell(row=row, column=1, value="Spending by Category").font = SUBHEAD_FONT
    row += 1
    headers = ["Category", "Monthly Avg"]
    rows = [[k, money(v)] for k, v in by_cat.items()]
    _, row = write_table(ws, row, 1, headers, rows)
    row += 2

    streams = cf.get("income_streams") or []
    ws.cell(row=row, column=1, value="Income Streams").font = SUBHEAD_FONT
    row += 1
    headers = ["Name", "Monthly Avg", "Volatility"]
    rows = [[s.get("name"), money(s.get("monthly_avg")), s.get("volatility")] for s in streams]
    write_table(ws, row, 1, headers, rows)

    autosize(ws)
    return ws


def build_debt(wb, model):
    ws = wb.create_sheet("Debt")
    row = write_title(ws, "Debt Strategy", span=5)

    debt = safe_get(model, "analysis", "debt", default={}) or {}
    ws.cell(row=row, column=1, value="Total Debt").font = BIG_LABEL_FONT
    ws.cell(row=row, column=2, value=money(debt.get("total"))).number_format = CURRENCY_FMT
    row += 2

    items = debt.get("items") or []
    ws.cell(row=row, column=1, value="Liabilities").font = SUBHEAD_FONT
    row += 1
    headers = ["Account", "Balance", "APR", "Min Payment"]
    rows = [[i.get("account_id"), money(i.get("balance")), i.get("apr"), money(i.get("min_payment"))] for i in items]
    first, row = write_table(ws, row, 1, headers, rows)
    for r in range(first, row):
        c = ws.cell(row=r, column=3)
        if isinstance(c.value, (int, float)):
            c.number_format = PCT_FMT
    row += 2

    ws.cell(row=row, column=1, value="Avalanche vs. Snowball (static — recompute via refresh)").font = SUBHEAD_FONT
    row += 1
    ava = debt.get("avalanche", {}) or {}
    snow = debt.get("snowball", {}) or {}
    headers = ["Strategy", "Months to Debt-Free", "Total Interest Paid", "Payoff Order"]
    rows = [
        ["Avalanche (highest APR first)", ava.get("months_to_free"), money(ava.get("total_interest")),
         ", ".join(ava.get("order") or [])],
        ["Snowball (smallest balance first)", snow.get("months_to_free"), money(snow.get("total_interest")),
         ", ".join(snow.get("order") or [])],
    ]
    write_table(ws, row, 1, headers, rows)

    autosize(ws)
    return ws


def build_retirement(wb, model):
    ws = wb.create_sheet("Retirement")
    row = write_title(ws, "Retirement / Long-Horizon Projection", span=5)

    ret = safe_get(model, "analysis", "retirement", default={}) or {}
    scen = ret.get("scenarios", {}) or {}
    ws.cell(row=row, column=1, value="Scenarios (static — recompute via refresh)").font = SUBHEAD_FONT
    row += 1
    headers = ["Scenario", "Return (%)", "Value at Retirement"]
    rows = []
    for name in ("best", "base", "worst"):
        s = scen.get(name, {}) or {}
        rows.append([name.capitalize(), s.get("return_pct"), money(s.get("value_at_retirement"))])
    first, row = write_table(ws, row, 1, headers, rows)
    for r in range(first, row):
        c = ws.cell(row=r, column=2)
        if isinstance(c.value, (int, float)):
            c.number_format = PCT_FMT
    row += 2

    proj = ret.get("projection_series") or []
    ws.cell(row=row, column=1, value="Projection Series").font = SUBHEAD_FONT
    row += 1
    header_row = row
    headers = ["Age", "Best", "Base", "Worst"]
    rows = [[p.get("age"), money(p.get("best")), money(p.get("base")), money(p.get("worst"))] for p in proj]
    first, row = write_table(ws, header_row, 1, headers, rows)

    if len(rows) > 1:
        line = LineChart()
        line.title = "Retirement Scenarios"
        line.y_axis.title = "Projected value ($)"
        line.x_axis.title = "Age"
        data = Reference(ws, min_col=2, max_col=4, min_row=header_row, max_row=row - 1)
        cats = Reference(ws, min_col=1, min_row=first, max_row=row - 1)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.width = 16
        line.height = 8
        ws.add_chart(line, f"F{header_row}")

    autosize(ws)
    return ws


def build_transactions(wb, model):
    ws = wb.create_sheet("Transactions")
    row = write_title(ws, "Transactions", span=6)

    txns = model.get("transactions", []) or []
    headers = ["Date", "Account", "Amount", "Description", "Category", "Income?", "Transfer?"]
    rows = [
        [t.get("date"), t.get("account_id"), money(t.get("amount")), t.get("description"),
         t.get("category"), "Yes" if t.get("is_income") else "No", "Yes" if t.get("is_transfer") else "No"]
        for t in txns
    ]
    if not rows:
        ws.cell(row=row, column=1, value="No transactions in plan model.")
        row += 1
    else:
        write_table(ws, row, 1, headers, rows)

    autosize(ws)
    return ws


def build_source_manifest(wb, model):
    ws = wb.create_sheet("Source Manifest")
    row = write_title(ws, "Source Manifest", span=5)
    ws.cell(row=row, column=1, value=f"Generated: {model.get('generated_at', 'unknown')}")
    ws.cell(row=row, column=1).font = Font(italic=True, color="6B7280")
    row += 2

    manifest = model.get("source_manifest", []) or []
    headers = ["Item", "Source", "As Of", "Confidence"]
    rows = [[m.get("item"), m.get("source"), m.get("as_of"), m.get("confidence")] for m in manifest]
    if not rows:
        ws.cell(row=row, column=1, value="No source manifest entries provided.")
        row += 1
    else:
        _, row = write_table(ws, row, 1, headers, rows)
    row += 2

    write_disclaimer(ws, model.get("disclaimer", ""), row)
    autosize(ws)
    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def render(model: dict, out_path: Path):
    wb = Workbook()
    build_dashboard(wb, model)
    build_assumptions(wb, model)
    build_net_worth(wb, model)
    build_cash_flow(wb, model)
    build_debt(wb, model)
    build_retirement(wb, model)
    build_transactions(wb, model)
    build_source_manifest(wb, model)
    wb.save(str(out_path))


def main():
    parser = argparse.ArgumentParser(description="Render plan_model.json to an interactive Excel workbook.")
    parser.add_argument("model_path", help="Path to plan_model.json")
    parser.add_argument("--out", default="financial_plan.xlsx", help="Output .xlsx path")
    args = parser.parse_args()

    model_path = Path(args.model_path)
    if not model_path.exists():
        print(f"Model file not found: {model_path}", file=sys.stderr)
        sys.exit(1)

    with open(model_path, "r", encoding="utf-8") as f:
        model = json.load(f)

    out_path = Path(args.out)
    render(model, out_path)
    print(f"Wrote {out_path.resolve()}")


if __name__ == "__main__":
    main()
